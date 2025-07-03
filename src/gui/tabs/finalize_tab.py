"""Processed documents tab for viewing and exporting reviewed documents."""

import csv
import json
import shutil
from collections.abc import Callable
from dataclasses import dataclass
from datetime import (
    datetime,
    timezone,
)  # Using timezone.utc for Python 3.10 compatibility (UTC added in 3.11)
from pathlib import Path

from PyQt6.QtCore import Qt, pyqtSignal
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import (
    QCheckBox,
    QComboBox,
    QGroupBox,
    QHBoxLayout,
    QHeaderView,
    QLabel,
    QLineEdit,
    QMessageBox,
    QSizePolicy,
    QSplitter,
    QTableWidget,
    QTableWidgetItem,
    QVBoxLayout,
    QWidget,
)

from src.constants import (
    DECISION_PANEL_MAX_HEIGHT,
    DEFAULT_PROCESSING_TIME,
    EXPORT_GROUP_MAX_HEIGHT,
    FINALIZE_SPLITTER_SIZES,
    FLAG_EMOJI,
    MAIN_LAYOUT_MARGINS,
    SEARCH_INPUT_MAX_WIDTH,
    STATS_LABEL_MAX_HEIGHT,
    TABLE_CHECKBOX_COLUMN_WIDTH,
    TABLE_FLAG_COLUMN_WIDTH,
)
from src.gui.styles import (
    create_primary_button,
    create_secondary_button,
    create_title_label,
)
from src.gui.widgets.document_viewer import DocumentViewer
from src.models.document import Document
from src.processing.document_store import DocumentStore
from src.processing.request_manager import RequestManager
from src.utils.statistics import calculate_document_statistics


@dataclass
class ProcessedDocument:
    """Document with review metadata."""

    document: Document
    review_timestamp: datetime
    processing_time: float  # seconds
    flagged_for_review: bool = False


class FinalizeTab(QWidget):
    """Tab for finalizing document review and generating export packages."""

    # Signals
    export_requested = pyqtSignal(list)  # List of documents to export
    package_requested = pyqtSignal(list)  # List of documents for FOIA package

    def __init__(
        self,
        request_manager: RequestManager | None = None,
        document_store: DocumentStore | None = None,
        audit_manager: "AuditManager | None" = None,
    ) -> None:
        """Initialize the processed tab."""
        super().__init__()
        self.request_manager = request_manager
        self.document_store = document_store
        self.audit_manager = audit_manager
        self.processed_documents: list[ProcessedDocument] = []
        self.filtered_documents: list[ProcessedDocument] = []
        self.source_folder: Path | None = None
        self._init_ui()

    def _init_ui(self) -> None:
        """Set up the user interface."""
        layout = QVBoxLayout()
        layout.setContentsMargins(*MAIN_LAYOUT_MARGINS)
        layout.setSpacing(10)

        # Header with title and request info
        header_layout = QHBoxLayout()

        # Title
        title = create_title_label("Finalize Package")
        header_layout.addWidget(title)

        header_layout.addStretch()

        # Active request info (top-right)
        if self.request_manager:
            self._active_request_label = QLabel("No active request")
            self._active_request_label.setStyleSheet(
                "font-size: 14px; color: #0066cc; font-weight: bold;"
            )
            header_layout.addWidget(self._active_request_label)
            self._update_active_request_display()

        layout.addLayout(header_layout)

        # Toolbar
        toolbar_layout = self._create_toolbar()
        layout.addLayout(toolbar_layout, 0)  # No stretch

        # Statistics bar
        self.stats_label = QLabel("Total: 0 | R: 0 | N: 0 | U: 0 | Agreement: 0%")
        self.stats_label.setStyleSheet(
            """
            QLabel {
                background-color: #f0f0f0;
                padding: 5px 10px;
                border-radius: 4px;
                font-weight: bold;
            }
            """
        )
        self.stats_label.setMaximumHeight(STATS_LABEL_MAX_HEIGHT)
        self.stats_label.setSizePolicy(
            QSizePolicy.Policy.Preferred, QSizePolicy.Policy.Maximum
        )
        layout.addWidget(self.stats_label)

        # Main content area with splitter
        self.splitter = QSplitter(Qt.Orientation.Horizontal)
        self.splitter.setHandleWidth(5)

        # Left panel - Document list
        left_panel = self._create_left_panel()
        self.splitter.addWidget(left_panel)

        # Right panel - Document viewer
        right_panel = self._create_right_panel()
        self.splitter.addWidget(right_panel)

        # Set initial splitter sizes (60/40 ratio)
        self.splitter.setSizes(FINALIZE_SPLITTER_SIZES)

        # Add splitter with stretch to take remaining space
        layout.addWidget(self.splitter, 1)  # Stretch factor of 1
        self.setLayout(layout)

        # Initially disable buttons (enabled when documents are added)
        self.update_button_states()

    def _create_toolbar(self) -> QHBoxLayout:
        """Create the toolbar with search and filter controls."""
        toolbar = QHBoxLayout()

        # Search input
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Search documents...")
        self.search_input.setMaximumWidth(SEARCH_INPUT_MAX_WIDTH)
        self.search_input.textChanged.connect(self.apply_filters)
        toolbar.addWidget(self.search_input)

        # Filter dropdown
        self.filter_dropdown = QComboBox()
        self.filter_dropdown.addItems(
            [
                "All Documents",
                "Responsive",
                "Non-Responsive",
                "Uncertain",
                "Disagreements Only",
            ]
        )
        self.filter_dropdown.currentIndexChanged.connect(self.apply_filters)
        toolbar.addWidget(self.filter_dropdown)

        toolbar.addStretch()

        # Select All Non-Duplicates button
        self.select_non_duplicates_btn = create_secondary_button("Select All Non-Duplicates")
        self.select_non_duplicates_btn.clicked.connect(self._select_non_duplicates)
        toolbar.addWidget(self.select_non_duplicates_btn)

        # Export button (dynamic text)
        self.export_button = create_secondary_button("Export (0)")
        self.export_button.clicked.connect(self.export_documents)
        toolbar.addWidget(self.export_button)

        # Generate package button
        self.generate_package_button = create_primary_button("Generate FOIA Package")
        self.generate_package_button.clicked.connect(self.generate_foia_package)
        toolbar.addWidget(self.generate_package_button)

        return toolbar

    def _create_left_panel(self) -> QWidget:
        """Create the left panel with document list and export options."""
        panel = QWidget()
        # Set size policy to allow expansion
        panel.setSizePolicy(QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding)

        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        # Document table
        self.document_table = QTableWidget()
        self.document_table.setColumnCount(7)
        self.document_table.setHorizontalHeaderLabels(
            ["", "Filename", "AI", "Human", "Duplicate Status", "Time", "Flag"]
        )
        self.document_table.setColumnWidth(
            0, TABLE_CHECKBOX_COLUMN_WIDTH
        )  # Checkbox column
        self.document_table.setColumnWidth(6, TABLE_FLAG_COLUMN_WIDTH)  # Flag column
        self.document_table.setSortingEnabled(False)  # Disable to preserve duplicate grouping
        self.document_table.setAlternatingRowColors(True)
        self.document_table.setSelectionBehavior(
            QTableWidget.SelectionBehavior.SelectRows
        )
        self.document_table.setSelectionMode(QTableWidget.SelectionMode.SingleSelection)
        self.document_table.itemSelectionChanged.connect(self.on_document_selected)
        # Set size policy for the table to expand
        self.document_table.setSizePolicy(
            QSizePolicy.Policy.Expanding, QSizePolicy.Policy.Expanding
        )
        # Make the filename column stretch to fill available space
        self.document_table.horizontalHeader().setStretchLastSection(False)
        self.document_table.horizontalHeader().setSectionResizeMode(
            1, QHeaderView.ResizeMode.Stretch
        )
        layout.addWidget(self.document_table, 1)  # Add stretch factor of 1

        # Export format options
        export_group = QGroupBox("Export Formats")
        export_group.setMaximumHeight(
            EXPORT_GROUP_MAX_HEIGHT
        )  # Set maximum height for export group
        export_layout = QHBoxLayout()

        self.csv_checkbox = QCheckBox("CSV")
        self.csv_checkbox.setChecked(True)
        self.excel_checkbox = QCheckBox("Excel")
        self.pdf_checkbox = QCheckBox("PDF")
        self.json_checkbox = QCheckBox("JSON")

        export_layout.addWidget(self.csv_checkbox)
        export_layout.addWidget(self.excel_checkbox)
        export_layout.addWidget(self.pdf_checkbox)
        export_layout.addWidget(self.json_checkbox)
        export_layout.addStretch()

        export_group.setLayout(export_layout)
        layout.addWidget(export_group, 0)  # No stretch for export group

        panel.setLayout(layout)
        return panel

    def _create_right_panel(self) -> QWidget:
        """Create the right panel with document viewer."""
        panel = QWidget()
        layout = QVBoxLayout()
        layout.setContentsMargins(0, 0, 0, 0)

        # Document viewer
        self.document_viewer = DocumentViewer()
        layout.addWidget(self.document_viewer)

        # Decision info panel
        self.decision_panel = QWidget()
        decision_layout = QVBoxLayout()

        self.ai_decision_label = QLabel("AI Classification: -")
        self.human_decision_label = QLabel("Human Decision: -")
        self.confidence_label = QLabel("Confidence: -")
        self.feedback_label = QLabel("Feedback: -")

        decision_layout.addWidget(self.ai_decision_label)
        decision_layout.addWidget(self.human_decision_label)
        decision_layout.addWidget(self.confidence_label)
        decision_layout.addWidget(self.feedback_label)

        # Add spacing before flag button
        decision_layout.addSpacing(15)

        # Flag for review button
        self.flag_button = create_secondary_button("Flag for Review")
        self.flag_button.clicked.connect(self.toggle_flag)
        self.flag_button.setEnabled(False)  # Initially disabled
        decision_layout.addWidget(self.flag_button)

        self.decision_panel.setLayout(decision_layout)
        self.decision_panel.setMaximumHeight(DECISION_PANEL_MAX_HEIGHT)
        layout.addWidget(self.decision_panel)

        panel.setLayout(layout)
        return panel

    def add_processed_document(self, document: Document) -> None:
        """Add a newly processed document from the review tab."""
        # Only add if it belongs to the active request
        if self.request_manager and self.document_store:
            active_request = self.request_manager.get_active_request()
            if active_request:
                # Check if document belongs to active request
                stored_doc = self.document_store.get_document(
                    active_request.id, document.filename
                )
                if stored_doc:
                    # For now, use a dummy processing time
                    # In a real implementation, this would be tracked during processing
                    processed_doc = ProcessedDocument(
                        document=document,
                        review_timestamp=datetime.now(
                            timezone.utc
                        ),  # timezone.utc for Python 3.10 compat
                        processing_time=DEFAULT_PROCESSING_TIME,  # Placeholder
                        flagged_for_review=False,
                    )
                    self.processed_documents.append(processed_doc)
                    self.apply_filters()
                    self.update_statistics()
                    self.update_button_states()
        else:
            # Fallback for when managers aren't available
            processed_doc = ProcessedDocument(
                document=document,
                review_timestamp=datetime.now(timezone.utc),
                processing_time=DEFAULT_PROCESSING_TIME,
                flagged_for_review=False,
            )
            self.processed_documents.append(processed_doc)
            self.apply_filters()
            self.update_statistics()
            self.update_button_states()

    def apply_filters(self) -> None:
        """Apply current filters to the document list."""
        # Define filter strategies
        filter_strategies: dict[int, Callable[[ProcessedDocument], bool]] = {
            0: lambda doc: True,  # All Documents
            1: lambda doc: doc.document.human_decision == "responsive",
            2: lambda doc: doc.document.human_decision == "non_responsive",
            3: lambda doc: doc.document.human_decision == "uncertain",
            4: lambda doc: doc.document.classification
            != doc.document.human_decision,  # Disagreements
        }

        # Start with all documents
        filtered = self.processed_documents.copy()

        # Apply text search
        search_text = self.search_input.text().lower()
        if search_text:
            filtered = [
                doc
                for doc in filtered
                if search_text in doc.document.filename.lower()
                or search_text in doc.document.content.lower()
            ]

        # Apply classification filter
        filter_index = self.filter_dropdown.currentIndex()
        filter_func = filter_strategies.get(filter_index, lambda doc: True)
        filtered = [doc for doc in filtered if filter_func(doc)]

        # Group documents with their duplicates
        self.filtered_documents = self._group_documents_with_duplicates(filtered)
        self.refresh_table()

    def _group_documents_with_duplicates(self, documents: list[ProcessedDocument]) -> list[ProcessedDocument]:
        """Group documents so duplicates appear immediately after their originals."""
        # Create a mapping of filenames to documents
        doc_map = {doc.document.filename: doc for doc in documents}
        
        # Track which documents we've already added to avoid duplicates
        added = set()
        grouped = []
        
        # First pass: Add originals and their duplicates
        for doc in documents:
            if doc.document.filename in added:
                continue
                
            # Add the document
            grouped.append(doc)
            added.add(doc.document.filename)
            
            # If this is an original document, find and add all its duplicates
            if not doc.document.is_duplicate:
                # Find all documents that are duplicates of this one
                duplicates = [
                    d for d in documents 
                    if d.document.is_duplicate 
                    and d.document.duplicate_of == doc.document.filename
                    and d.document.filename not in added
                ]
                # Sort duplicates by similarity score (highest first)
                duplicates.sort(
                    key=lambda x: x.document.similarity_score or 0, 
                    reverse=True
                )
                # Add duplicates right after the original
                for dup in duplicates:
                    grouped.append(dup)
                    added.add(dup.document.filename)
        
        return grouped

    def refresh_table(self) -> None:
        """Refresh the document table with filtered results."""
        self.document_table.setRowCount(len(self.filtered_documents))

        for row, proc_doc in enumerate(self.filtered_documents):
            doc = proc_doc.document

            # Checkbox with padding (unchecked by default for duplicates)
            checkbox_widget = QWidget()
            checkbox_layout = QHBoxLayout(checkbox_widget)
            checkbox_layout.setContentsMargins(8, 0, 0, 0)  # Add 8px left padding
            checkbox = QCheckBox()
            checkbox.setChecked(not doc.is_duplicate)  # Uncheck duplicates by default
            checkbox.stateChanged.connect(self.update_export_button)
            checkbox_layout.addWidget(checkbox)
            checkbox_layout.addStretch()
            self.document_table.setCellWidget(row, 0, checkbox_widget)

            # Filename (with indentation for duplicates)
            filename_text = "    ↳ " + doc.filename if doc.is_duplicate else doc.filename
            filename_item = QTableWidgetItem(filename_text)
            if doc.is_duplicate:
                filename_item.setForeground(QColor("#666666"))  # Gray text for duplicates
            self.document_table.setItem(row, 1, filename_item)

            # AI Classification
            ai_item = QTableWidgetItem(doc.classification or "-")
            self.document_table.setItem(row, 2, ai_item)

            # Human Decision
            human_item = QTableWidgetItem(doc.human_decision or "-")
            self.document_table.setItem(row, 3, human_item)

            # Duplicate Status
            if doc.is_duplicate:
                if doc.similarity_score and doc.similarity_score < 1.0:
                    status_text = f"{doc.similarity_score:.0%} similar"
                else:
                    status_text = "Exact duplicate"
                status_item = QTableWidgetItem(status_text)
                status_item.setForeground(QColor("#666666"))  # Gray text
            else:
                # Check if this original has any duplicates
                has_duplicates = any(
                    d.document.is_duplicate and d.document.duplicate_of == doc.filename
                    for d in self.filtered_documents
                )
                status_text = "Original (has duplicates)" if has_duplicates else "Original"
                status_item = QTableWidgetItem(status_text)
                if has_duplicates:
                    status_item.setForeground(QColor("#0066cc"))  # Blue text for originals with duplicates
            self.document_table.setItem(row, 4, status_item)

            # Processing Time
            time_item = QTableWidgetItem(f"{proc_doc.processing_time:.1f}s")
            self.document_table.setItem(row, 5, time_item)

            # Flag
            flag_item = QTableWidgetItem(
                FLAG_EMOJI if proc_doc.flagged_for_review else ""
            )
            flag_item.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
            self.document_table.setItem(row, 6, flag_item)

            # Style row based on status
            if doc.is_duplicate:
                # Light gray background for duplicate rows
                for col in range(1, self.document_table.columnCount()):
                    item = self.document_table.item(row, col)
                    if item:
                        item.setBackground(QColor(245, 245, 245))
            elif doc.classification != doc.human_decision:
                # Pink background for disagreements
                for col in range(1, self.document_table.columnCount()):
                    item = self.document_table.item(row, col)
                    if item:
                        item.setBackground(QColor(255, 200, 200))
        
        # Update export button text to reflect selected count
        self.update_export_button()

    def on_document_selected(self) -> None:
        """Handle document selection in the table."""
        current_row = self.document_table.currentRow()
        if 0 <= current_row < len(self.filtered_documents):
            proc_doc = self.filtered_documents[current_row]
            self.display_document(proc_doc)

    def display_document(self, proc_doc: ProcessedDocument) -> None:
        """Display the selected document in the viewer."""
        doc = proc_doc.document
        self.document_viewer.display_document(doc.filename, doc.content, doc.exemptions)

        # Log document view to audit trail
        if self.audit_manager and self.request_manager:
            active_request = self.request_manager.get_active_request()
            if active_request:
                self.audit_manager.log_view(
                    filename=doc.filename,
                    tab_name="Finalize",
                    request_id=active_request.id
                )

        # Update decision info
        self.ai_decision_label.setText(
            f"AI Classification: {doc.classification or '-'}"
        )
        self.human_decision_label.setText(
            f"Human Decision: {doc.human_decision or '-'}"
        )
        confidence_text = (
            f"Confidence: {doc.confidence:.2f}" if doc.confidence else "Confidence: -"
        )
        self.confidence_label.setText(confidence_text)

        # Update feedback label
        if doc.human_feedback:
            self.feedback_label.setText(f"Feedback: {doc.human_feedback}")
        else:
            self.feedback_label.setText("Feedback: -")

        # Update flag button
        self.flag_button.setEnabled(True)  # Enable when document is displayed
        if proc_doc.flagged_for_review:
            self.flag_button.setText("Remove Flag")
        else:
            self.flag_button.setText("Flag for Review")

    def update_statistics(self) -> None:
        """Update the statistics display."""
        stats = calculate_document_statistics(self.processed_documents)
        self.stats_label.setText(stats.to_display_string())

    def update_export_button(self) -> None:
        """Update export button text based on selection."""
        selected_count = 0
        for row in range(self.document_table.rowCount()):
            widget = self.document_table.cellWidget(row, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    selected_count += 1

        if selected_count > 0:
            self.export_button.setText(f"Export ({selected_count})")
        else:
            self.export_button.setText("Export (0)")

    def update_button_states(self) -> None:
        """Update the enabled state of export and FOIA package buttons."""
        # Enable buttons if at least one document has been reviewed
        has_reviewed_documents = len(self.processed_documents) > 0

        self.export_button.setEnabled(has_reviewed_documents)
        self.generate_package_button.setEnabled(has_reviewed_documents)
        self.select_non_duplicates_btn.setEnabled(has_reviewed_documents)

    def _select_non_duplicates(self) -> None:
        """Select all non-duplicate documents in the table."""
        for row in range(self.document_table.rowCount()):
            if row < len(self.filtered_documents):
                doc = self.filtered_documents[row].document
                widget = self.document_table.cellWidget(row, 0)
                if widget:
                    checkbox = widget.findChild(QCheckBox)
                    if checkbox:
                        checkbox.setChecked(not doc.is_duplicate)

        # Update the export button text
        self.update_export_button()

    def get_selected_documents(self) -> list[ProcessedDocument]:
        """Get list of selected documents."""
        selected = []
        for row in range(self.document_table.rowCount()):
            widget = self.document_table.cellWidget(row, 0)
            if widget:
                checkbox = widget.findChild(QCheckBox)
                if checkbox and checkbox.isChecked():
                    selected.append(self.filtered_documents[row])
        return selected

    def set_source_folder(self, folder: Path) -> None:
        """Set the source folder for FOIA package generation."""
        self.source_folder = folder

    def set_all_documents_reviewed(self, reviewed: bool) -> None:
        """Called when all documents have been reviewed (for compatibility).

        Note: Buttons are now enabled when at least one document is reviewed,
        not when all documents are reviewed.
        """
        self.update_button_states()

    def clear_all(self) -> None:
        """Clear all documents from the finalize tab."""
        self.processed_documents.clear()
        self.filtered_documents.clear()
        self.document_table.setRowCount(0)
        self.document_viewer.clear()
        self.ai_decision_label.setText("AI Classification: -")
        self.human_decision_label.setText("Human Decision: -")
        self.confidence_label.setText("Confidence: -")
        self.feedback_label.setText("Feedback: -")
        self.flag_button.setEnabled(False)  # Disable when clearing
        self.flag_button.setText("Flag for Review")  # Reset text
        self.update_statistics()
        self.update_export_button()
        self.update_button_states()

    def toggle_flag(self) -> None:
        """Toggle flag for review on current document."""
        current_row = self.document_table.currentRow()
        if 0 <= current_row < len(self.filtered_documents):
            proc_doc = self.filtered_documents[current_row]
            proc_doc.flagged_for_review = not proc_doc.flagged_for_review
            self.refresh_table()
            self.display_document(proc_doc)

    def export_documents(self) -> None:
        """Export documents in selected formats."""
        # Check if source folder is set
        if not self.source_folder:
            QMessageBox.warning(
                self,
                "No Source Folder",
                "Please select a folder in the Intake tab first.",
            )
            return

        # Get documents to export
        selected = self.get_selected_documents()
        if not selected and not self.processed_documents:
            QMessageBox.warning(self, "No Documents", "No documents to export.")
            return
        
        # If there are documents but none are selected, show a different message
        if not selected and self.processed_documents:
            QMessageBox.warning(
                self, 
                "No Documents Selected", 
                "No documents are selected for export. Please select at least one document."
            )
            return

        documents_to_export = selected

        try:
            # Create export directory in Documents folder
            export_base_dir = Path.home() / "Documents" / "FOIA_Exports"
            export_base_dir.mkdir(exist_ok=True)

            # Create timestamped subdirectory and remove if exists
            timestamp = datetime.now(timezone.utc).strftime(
                "%Y%m%d_%H%M%S"
            )  # timezone.utc for Python 3.10 compat
            export_dir = export_base_dir / f"Export_{timestamp}"

            if export_dir.exists():
                shutil.rmtree(export_dir)
            export_dir.mkdir(exist_ok=True)

            # Export in selected formats
            exported_files = []

            if self.csv_checkbox.isChecked():
                filename = self._export_csv(
                    documents_to_export, str(export_dir), timestamp
                )
                if filename:
                    exported_files.append(Path(filename).name)

            if self.json_checkbox.isChecked():
                filename = self._export_json(
                    documents_to_export, str(export_dir), timestamp
                )
                if filename:
                    exported_files.append(Path(filename).name)

            if self.excel_checkbox.isChecked():
                filename = self._export_excel(
                    documents_to_export, str(export_dir), timestamp
                )
                if filename:
                    exported_files.append(Path(filename).name)

            if self.pdf_checkbox.isChecked():
                filename = self._export_pdf(
                    documents_to_export, str(export_dir), timestamp
                )
                if filename:
                    exported_files.append(Path(filename).name)

            # Log export to audit trail
            if exported_files and self.audit_manager and self.request_manager:
                active_request = self.request_manager.get_active_request()
                if active_request:
                    formats = []
                    if self.csv_checkbox.isChecked():
                        formats.append("CSV")
                    if self.json_checkbox.isChecked():
                        formats.append("JSON")
                    if self.excel_checkbox.isChecked():
                        formats.append("Excel")
                    if self.pdf_checkbox.isChecked():
                        formats.append("PDF")
                    
                    format_str = ", ".join(formats)
                    self.audit_manager.log_export(
                        format=format_str,
                        document_count=len(documents_to_export),
                        request_id=active_request.id
                    )

            # Show success message with option to open folder
            if exported_files:
                msg_box = QMessageBox(self)
                msg_box.setWindowTitle("Export Complete")
                msg_box.setText(
                    f"Exported {len(documents_to_export)} documents successfully!\n\n"
                    f"Location: {export_dir}\n"
                    f"Files: {', '.join(exported_files)}"
                )
                msg_box.setStandardButtons(
                    QMessageBox.StandardButton.Open | QMessageBox.StandardButton.Ok
                )
                msg_box.setDefaultButton(QMessageBox.StandardButton.Open)

                result = msg_box.exec()
                if result == QMessageBox.StandardButton.Open:
                    # Open the folder in the system file manager
                    import platform
                    import subprocess

                    if platform.system() == "Windows":
                        subprocess.run(["explorer", str(export_dir)])
                    elif platform.system() == "Darwin":  # macOS
                        subprocess.run(["open", str(export_dir)])
                    else:  # Linux and others
                        subprocess.run(["xdg-open", str(export_dir)])

        except Exception as e:
            QMessageBox.critical(
                self,
                "Export Error",
                f"Failed to export documents: {e}",
            )

    def _export_csv(
        self, documents: list[ProcessedDocument], export_dir: str, timestamp: str
    ) -> str | None:
        """Export documents to CSV format."""
        try:
            filepath = Path(export_dir) / f"foia_export_{timestamp}.csv"
            with open(filepath, "w", newline="", encoding="utf-8") as csvfile:
                fieldnames = [
                    "filename",
                    "ai_classification",
                    "human_decision",
                    "confidence",
                    "review_timestamp",
                    "processing_time",
                    "flagged",
                    "justification",
                    "exemption_count",
                ]
                writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                writer.writeheader()

                for proc_doc in documents:
                    doc = proc_doc.document
                    writer.writerow(
                        {
                            "filename": doc.filename,
                            "ai_classification": doc.classification or "",
                            "human_decision": doc.human_decision or "",
                            "confidence": (
                                f"{doc.confidence:.2f}" if doc.confidence else ""
                            ),
                            "review_timestamp": proc_doc.review_timestamp.isoformat(),
                            "processing_time": f"{proc_doc.processing_time:.1f}",
                            "flagged": "Yes" if proc_doc.flagged_for_review else "No",
                            "justification": doc.justification or "",
                            "exemption_count": (
                                len(doc.exemptions) if doc.exemptions else 0
                            ),
                        }
                    )
            return str(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export CSV: {e}")
            return None

    def _export_json(
        self, documents: list[ProcessedDocument], export_dir: str, timestamp: str
    ) -> str | None:
        """Export documents to JSON format."""
        try:
            filepath = Path(export_dir) / f"foia_export_{timestamp}.json"
            data = []

            for proc_doc in documents:
                doc = proc_doc.document
                data.append(
                    {
                        "filename": doc.filename,
                        "content": doc.content,
                        "ai_classification": doc.classification,
                        "human_decision": doc.human_decision,
                        "confidence": doc.confidence,
                        "justification": doc.justification,
                        "exemptions": doc.exemptions or [],
                        "human_feedback": doc.human_feedback,
                        "review_timestamp": proc_doc.review_timestamp.isoformat(),
                        "processing_time": proc_doc.processing_time,
                        "flagged_for_review": proc_doc.flagged_for_review,
                    }
                )

            with open(filepath, "w", encoding="utf-8") as f:
                json.dump(data, f, indent=2, ensure_ascii=False)

            return str(filepath)
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export JSON: {e}")
            return None

    def _export_excel(
        self, documents: list[ProcessedDocument], export_dir: str, timestamp: str
    ) -> str | None:
        """Export documents to Excel format."""
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.utils import get_column_letter

            filepath = Path(export_dir) / f"foia_export_{timestamp}.xlsx"

            # Create workbook and get active sheet
            wb = Workbook()
            ws = wb.active
            ws.title = "FOIA Documents"

            # Define headers
            headers = [
                "Filename",
                "AI Classification",
                "Human Decision",
                "Agreement",
                "Confidence",
                "Review Timestamp",
                "Processing Time (s)",
                "Flagged",
                "Justification",
                "Exemptions",
                "Human Feedback"
            ]

            # Style for headers
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")

            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Style for disagreement rows
            disagreement_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

            # Write data
            for row_idx, proc_doc in enumerate(documents, 2):
                doc = proc_doc.document

                # Check if there's disagreement
                has_disagreement = doc.classification != doc.human_decision

                # Write data
                row_data = [
                    doc.filename,
                    doc.classification or "",
                    doc.human_decision or "",
                    "No" if has_disagreement else "Yes",
                    f"{doc.confidence:.2f}" if doc.confidence else "",
                    proc_doc.review_timestamp.isoformat(),
                    f"{proc_doc.processing_time:.1f}",
                    "Yes" if proc_doc.flagged_for_review else "No",
                    doc.justification or "",
                    len(doc.exemptions) if doc.exemptions else 0,
                    doc.human_feedback or ""
                ]

                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)

                    # Highlight disagreement rows
                    if has_disagreement:
                        cell.fill = disagreement_fill

            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width

            # Add summary sheet
            summary_ws = wb.create_sheet("Summary")

            # Calculate statistics
            total_docs = len(documents)
            responsive = sum(1 for d in documents if d.document.human_decision == "responsive")
            non_responsive = sum(1 for d in documents if d.document.human_decision == "non_responsive")
            uncertain = sum(1 for d in documents if d.document.human_decision == "uncertain")
            agreements = sum(1 for d in documents if d.document.classification == d.document.human_decision)
            flagged = sum(1 for d in documents if d.flagged_for_review)

            # Write summary
            summary_data = [
                ["FOIA Processing Summary", ""],
                ["", ""],
                ["Total Documents", total_docs],
                ["Responsive", responsive],
                ["Non-Responsive", non_responsive],
                ["Uncertain", uncertain],
                ["", ""],
                ["AI/Human Agreement", f"{(agreements/total_docs*100):.1f}%" if total_docs > 0 else "0%"],
                ["Documents Flagged", flagged],
                ["", ""],
                ["Export Date", datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")]
            ]

            for row_idx, (label, value) in enumerate(summary_data, 1):
                summary_ws.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
                summary_ws.cell(row=row_idx, column=2, value=value)

            # Adjust summary column widths
            summary_ws.column_dimensions['A'].width = 25
            summary_ws.column_dimensions['B'].width = 20

            # Save workbook
            wb.save(filepath)

            return str(filepath)

        except ImportError:
            QMessageBox.critical(
                self,
                "Export Error",
                "openpyxl is not installed. Please install it with: pip install openpyxl"
            )
            return None
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export Excel: {e}")
            return None

    def _export_pdf(
        self, documents: list[ProcessedDocument], export_dir: str, timestamp: str
    ) -> str | None:
        """Export documents to PDF format."""
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import landscape, letter
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )
            from reportlab.platypus.tableofcontents import TableOfContents

            filepath = Path(export_dir) / f"foia_export_{timestamp}.pdf"

            # Create PDF document
            pdf_doc = SimpleDocTemplate(
                str(filepath),
                pagesize=landscape(letter),
                rightMargin=0.5*inch,
                leftMargin=0.5*inch,
                topMargin=0.5*inch,
                bottomMargin=0.5*inch
            )

            # Container for the 'Flowable' objects
            elements = []
            styles = getSampleStyleSheet()

            # Title style
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#366092'),
                spaceAfter=30,
                alignment=1  # Center alignment
            )

            # Add title
            elements.append(Paragraph("FOIA Document Export Report", title_style))
            elements.append(Spacer(1, 0.2*inch))

            # Add export date
            date_style = ParagraphStyle(
                'DateStyle',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            elements.append(
                Paragraph(
                    f"Generated: {datetime.now(timezone.utc).strftime('%B %d, %Y at %H:%M:%S UTC')}",
                    date_style
                )
            )
            elements.append(Spacer(1, 0.5*inch))

            # Summary section
            summary_style = ParagraphStyle(
                'SummaryHeading',
                parent=styles['Heading2'],
                fontSize=16,
                textColor=colors.HexColor('#366092'),
                spaceAfter=12
            )
            elements.append(Paragraph("Summary Statistics", summary_style))

            # Calculate statistics
            total_docs = len(documents)
            responsive = sum(1 for d in documents if d.document.human_decision == "responsive")
            non_responsive = sum(1 for d in documents if d.document.human_decision == "non_responsive")
            uncertain = sum(1 for d in documents if d.document.human_decision == "uncertain")
            agreements = sum(1 for d in documents if d.document.classification == d.document.human_decision)
            flagged = sum(1 for d in documents if d.flagged_for_review)

            # Create summary table
            summary_data = [
                ['Metric', 'Value'],
                ['Total Documents', str(total_docs)],
                ['Responsive', str(responsive)],
                ['Non-Responsive', str(non_responsive)],
                ['Uncertain', str(uncertain)],
                ['AI/Human Agreement', f"{(agreements/total_docs*100):.1f}%" if total_docs > 0 else "0%"],
                ['Documents Flagged', str(flagged)]
            ]

            summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
            summary_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
            ]))

            elements.append(summary_table)
            elements.append(PageBreak())

            # Document details section
            elements.append(Paragraph("Document Details", summary_style))
            elements.append(Spacer(1, 0.2*inch))

            # Create document table headers
            headers = [
                'Filename',
                'AI Class',
                'Human',
                'Match',
                'Conf',
                'Time',
                'Flag'
            ]

            # Prepare data for document table
            doc_data = [headers]

            for proc_doc in documents:
                doc = proc_doc.document
                has_disagreement = doc.classification != doc.human_decision

                row = [
                    doc.filename[:30] + "..." if len(doc.filename) > 30 else doc.filename,
                    doc.classification or "-",
                    doc.human_decision or "-",
                    "No" if has_disagreement else "Yes",
                    f"{doc.confidence:.2f}" if doc.confidence else "-",
                    f"{proc_doc.processing_time:.1f}s",
                    "🚩" if proc_doc.flagged_for_review else ""
                ]
                doc_data.append(row)

            # Create document table with adjusted column widths
            doc_table = Table(
                doc_data,
                colWidths=[3*inch, 1.3*inch, 1.3*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.5*inch]
            )

            # Define table style
            table_style = [
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#366092')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
                ('ALIGN', (0, 0), (0, -1), 'LEFT'),  # Left align filenames
            ]

            # Highlight disagreement rows
            for i, proc_doc in enumerate(documents, 1):
                if proc_doc.document.classification != proc_doc.document.human_decision:
                    table_style.append(('BACKGROUND', (0, i), (-1, i), colors.HexColor('#FFC7CE')))

            doc_table.setStyle(TableStyle(table_style))
            elements.append(doc_table)

            # Add page break before exemptions if needed
            if any(d.document.exemptions for d in documents):
                elements.append(PageBreak())
                elements.append(Paragraph("Exemption Details", summary_style))
                elements.append(Spacer(1, 0.2*inch))

                # Create exemption summary
                exemption_count = sum(len(d.document.exemptions) if d.document.exemptions else 0 for d in documents)
                elements.append(
                    Paragraph(
                        f"Total exemptions found: {exemption_count}",
                        styles['Normal']
                    )
                )
                elements.append(Spacer(1, 0.1*inch))

                # List documents with exemptions
                for proc_doc in documents:
                    if proc_doc.document.exemptions:
                        elements.append(
                            Paragraph(
                                f"• {proc_doc.document.filename}: {len(proc_doc.document.exemptions)} exemptions",
                                styles['Normal']
                            )
                        )

            # Build PDF
            pdf_doc.build(elements)

            return str(filepath)

        except ImportError:
            QMessageBox.critical(
                self,
                "Export Error",
                "reportlab is not installed. Please install it with: pip install reportlab"
            )
            return None
        except Exception as e:
            QMessageBox.critical(self, "Export Error", f"Failed to export PDF: {e}")
            return None

    def generate_foia_package(self) -> None:
        """Generate complete FOIA response package."""
        # Check if source folder is set
        if not self.source_folder:
            QMessageBox.warning(
                self,
                "No Source Folder",
                "Please select a folder in the Intake tab first.",
            )
            return

        # Get responsive documents only
        responsive_docs = [
            doc
            for doc in self.processed_documents
            if doc.document.human_decision == "responsive"
        ]

        if not responsive_docs:
            QMessageBox.warning(
                self,
                "No Responsive Documents",
                "No responsive documents found to include in FOIA package.",
            )
            return

        try:
            # Create package directory in Documents folder
            package_base_dir = Path.home() / "Documents" / "FOIA_Packages"
            package_base_dir.mkdir(exist_ok=True)

            # Get request name if available
            request_name = ""
            if self.request_manager:
                active_request = self.request_manager.get_active_request()
                if active_request and active_request.name:
                    # Sanitize request name for folder
                    request_name = "_" + "".join(c for c in active_request.name if c.isalnum() or c in (' ', '-', '_')).rstrip()
                    request_name = request_name.replace(' ', '_')

            # Create package directory (overwrite if exists)
            package_dir = package_base_dir / f"FOIA_Response{request_name}"

            if package_dir.exists():
                shutil.rmtree(package_dir)
            package_dir.mkdir(exist_ok=True)

            # Create subdirectories
            responsive_dir = package_dir / "responsive_documents"
            responsive_dir.mkdir(exist_ok=True)

            # Copy responsive documents
            for proc_doc in responsive_docs:
                doc = proc_doc.document
                doc_path = responsive_dir / doc.filename
                doc_path.write_text(doc.content, encoding="utf-8")

            # Generate exemption log
            self._generate_exemption_log(responsive_docs, package_dir)

            # Generate summary report
            self._generate_summary_report(package_dir)

            # Generate cover letter template
            self._generate_cover_letter(package_dir, len(responsive_docs))

            # Log FOIA package generation to audit trail
            if self.audit_manager and self.request_manager:
                active_request = self.request_manager.get_active_request()
                if active_request:
                    self.audit_manager.log_export(
                        format="FOIA Package",
                        document_count=len(responsive_docs),
                        request_id=active_request.id
                    )

            # Show success with option to open folder
            msg_box = QMessageBox(self)
            msg_box.setWindowTitle("Package Generated")
            msg_box.setText(
                f"FOIA response package generated successfully!\n\n"
                f"Location: {package_dir}\n"
                f"Included: {len(responsive_docs)} responsive documents"
            )
            msg_box.setStandardButtons(
                QMessageBox.StandardButton.Open | QMessageBox.StandardButton.Ok
            )
            msg_box.setDefaultButton(QMessageBox.StandardButton.Open)

            result = msg_box.exec()
            if result == QMessageBox.StandardButton.Open:
                # Open the folder in the system file manager
                import platform
                import subprocess

                if platform.system() == "Windows":
                    subprocess.run(["explorer", str(package_dir)])
                elif platform.system() == "Darwin":  # macOS
                    subprocess.run(["open", str(package_dir)])
                else:  # Linux and others
                    subprocess.run(["xdg-open", str(package_dir)])

        except Exception as e:
            QMessageBox.critical(
                self,
                "Package Generation Error",
                f"Failed to generate FOIA package: {e}",
            )

    def _generate_exemption_log(
        self, responsive_docs: list[ProcessedDocument], package_dir: Path
    ) -> None:
        """Generate CSV log of all exemptions applied."""
        exemption_log_path = package_dir / "exemption_log.csv"

        with open(exemption_log_path, "w", newline="", encoding="utf-8") as csvfile:
            fieldnames = [
                "filename",
                "exemption_type",
                "exemption_code",
                "redacted_text",
                "location",
            ]
            writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
            writer.writeheader()

            for proc_doc in responsive_docs:
                doc = proc_doc.document
                if doc.exemptions:
                    for exemption in doc.exemptions:
                        writer.writerow(
                            {
                                "filename": doc.filename,
                                "exemption_type": exemption.get("type", ""),
                                "exemption_code": exemption.get("exemption_code", ""),
                                "redacted_text": exemption.get("text", ""),
                                "location": f"{exemption.get('start', '')}-{exemption.get('end', '')}",
                            }
                        )

    def _generate_summary_report(self, package_dir: Path) -> None:
        """Generate processing summary report."""
        summary_path = package_dir / "processing_summary.txt"
        stats = calculate_document_statistics(self.processed_documents)

        with open(summary_path, "w", encoding="utf-8") as f:
            f.write("FOIA PROCESSING SUMMARY REPORT\n")
            f.write("=" * 50 + "\n\n")
            f.write(
                f"Generated: {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S')}\n\n"  # timezone.utc for Python 3.10 compat
            )
            f.write("DOCUMENT STATISTICS:\n")
            f.write(f"Total Documents Processed: {stats.total}\n")
            f.write(f"Responsive Documents: {stats.responsive}\n")
            f.write(f"Non-Responsive Documents: {stats.non_responsive}\n")
            f.write(f"Uncertain Documents: {stats.uncertain}\n\n")
            f.write(f"AI/Human Agreement Rate: {stats.agreement_rate:.1f}%\n\n")

            # List flagged documents
            flagged = [d for d in self.processed_documents if d.flagged_for_review]
            if flagged:
                f.write(f"DOCUMENTS FLAGGED FOR REVIEW ({len(flagged)}):\n")
                for doc in flagged:
                    f.write(f"- {doc.document.filename}\n")

    def _generate_cover_letter(self, package_dir: Path, responsive_count: int) -> None:
        """Generate FOIA response cover letter template."""
        cover_letter_path = package_dir / "cover_letter_template.txt"

        with open(cover_letter_path, "w", encoding="utf-8") as f:
            f.write("[AGENCY LETTERHEAD]\n\n")
            f.write(
                f"Date: {datetime.now(timezone.utc).strftime('%B %d, %Y')}\n\n"
            )  # timezone.utc for Python 3.10 compat
            f.write("[Requester Name]\n")
            f.write("[Requester Address]\n\n")
            f.write("Re: Freedom of Information Act Request\n\n")
            f.write("Dear [Requester Name]:\n\n")
            f.write(
                "This letter is in response to your Freedom of Information Act (FOIA) request "
            )
            f.write(
                "dated [DATE], in which you requested [DESCRIPTION OF REQUEST].\n\n"
            )
            f.write(
                "We have completed our search and review of records responsive to your request. "
            )
            f.write(
                f"We identified {responsive_count} document(s) that are responsive to your request.\n\n"
            )
            f.write(
                "The responsive documents are enclosed with this letter. Some information has been "
            )
            f.write("redacted pursuant to the following FOIA exemptions:\n\n")
            f.write(
                "- Exemption (b)(6): Information that would constitute a clearly unwarranted "
            )
            f.write("invasion of personal privacy\n\n")
            f.write(
                "If you have any questions regarding this response, please contact [CONTACT NAME] "
            )
            f.write("at [CONTACT INFO].\n\n")
            f.write("Sincerely,\n\n")
            f.write("[Name]\n")
            f.write("[Title]\n")
            f.write("[Agency]\n")

    def _update_active_request_display(self) -> None:
        """Update the active request display."""
        if hasattr(self, "_active_request_label") and self.request_manager:
            active_request = self.request_manager.get_active_request()
            if active_request:
                self._active_request_label.setText(f"Request: {active_request.name}")
            else:
                self._active_request_label.setText("No active request")

    def refresh_request_context(self) -> None:
        """Refresh documents for the active request."""
        self._update_active_request_display()

        # Clear current documents
        self.processed_documents.clear()

        # Load reviewed documents from document store for active request
        if self.document_store and self.request_manager:
            active_request = self.request_manager.get_active_request()
            if active_request:
                # Get all reviewed documents
                reviewed_docs = self.document_store.get_reviewed_documents(
                    active_request.id
                )

                # Convert to ProcessedDocument objects
                for doc in reviewed_docs:
                    processed_doc = ProcessedDocument(
                        document=doc,
                        review_timestamp=datetime.now(timezone.utc),
                        processing_time=DEFAULT_PROCESSING_TIME,
                        flagged_for_review=False,
                    )
                    self.processed_documents.append(processed_doc)

        # Update display
        self.apply_filters()
        self.update_statistics()
        self.update_button_states()
